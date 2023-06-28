import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook

# 시간 측정 시작
start_time = time.time()

excel_file = 'lis.xlsx'
workbook = load_workbook(filename=excel_file)
sheet = workbook['filename']
excel_item_data = []
item_ratio_data = ["묶음상품 비율(%)"]
# 맨 위 셀 제외
for cell in sheet['C'][1:]:
    excel_item_data.append(cell.value)


# options = webdriver.ChromeOptions()
# options.add_argument("headless")
# browser = webdriver.Chrome(options=options)
browser = webdriver.Chrome()
browser.maximize_window()

def search(item):
    url = "https://itemscout.io/"
    browser.get(url)
    # time.sleep(3)
    browser.implicitly_wait(10)
    # WebDriverWait(browser, 5).until(EC.element_to_be_clickable((By.XPATH, "//*[@id='explore']/div/form/div/input")))
    search_input = browser.find_element(By.XPATH, "//*[@id='explore']/div/form/div/input")
    search_input.send_keys(item)
    search_input.send_keys("\n")
    # time.sleep(3)
    browser.implicitly_wait(10)
    # WebDriverWait(browser, 5).until(EC.presence_of_element_located((By.XPATH, "//*[@id='app']/div/main/div/div/div[2]/div/div[1]/div[2]/div/div[2]/div[1]/div[2]/div[1]/div[2]/div[3]/div[2]")))
    item_ratio_text = browser.find_element(By.XPATH, "//*[@id='app']/div/main/div/div/div[2]/div/div[1]/div[2]/div/div[2]/div[1]/div[2]/div[1]/div[2]/div[3]/div[2]").text
    item_ratio_data.append(item_ratio_text[0:2])

for item in excel_item_data[0:3]:
    search(item)

for i, data in enumerate(item_ratio_data, start=1):
    cell = sheet.cell(row=i, column=1)
    cell.value = data


workbook.save('result.xlsx')
browser.quit()

# 시간 측정 종료
end_time = time.time()
execution_time = end_time - start_time
print(f"실행 시간: {execution_time}초")
